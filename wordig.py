# C:\>wordig.py -u 가①⑴ⓐ⒜ㄱ㉠㉮㈀㈎𐊀
from pathlib import Path
from pathlib import PurePath
import argparse
import bs4
import csv
import glob
import xml.etree.ElementTree as ET
import openpyxl
import os
import pymupdf
import re
import unicodedata

class WordDigger(object):

    # flag: ONCE, EXPAND

    def __init__(self, targets, options={}):

        self.targets = targets
        self.options = {
            'aim': None,
            'aim_pattern': None,
            'substitute': None,
            'pattern': None,
            'compare': False,
            'case_sensitive': True,
            'dotall': False,
            'flag': 'ONCE',
            'extract': False,
            'gather': False,
            'output': None,
            'overwrite': False,
            'recursive': False,
            'page_count': False,
            'unicode': False,
            'unicode_bits': False,
            'unicode_hexadecimal': False,
            'unicode_decimal': False,
            'xlsx': False,
            'tsv': False,
            'escape_tex': False,
            'quietly': False,
            'beautify': False,
            'indent': None
        }

        self.files = 0
        self.lines = 0
        self.spc_chars = 0
        self.chars = 0
        self.words = 0
        self.pages = 0
        self.found = []
        self.found_count = {}

        for key in self.options.keys():
            if key in options:
                self.options[key] = options.get(key)

        if self.options['flag'] == '1':
            self.options['flag'] == 'ONCE'
        if self.options['flag'] == '2':
            self.options['flag'] == 'EXPAND'

        self.determine_task()


    def run_recursive(self, func:str) -> None:

        if self.options['recursive']:
            for target in self.targets:
                dir = os.path.dirname(target)
                fileNamePattern = os.path.basename(target)
                if dir == '': dir = '.'
                subdirs = [x[0] for x in os.walk(dir)]
                for subdir in subdirs:
                    files = os.path.join(subdir, fileNamePattern).replace('/','\\')
                    for filePath in glob.glob(files):
                        func(filePath)
        else:
            for target in self.targets: 
                for filePath in glob.glob(target):
                    func(filePath)


    def find(self, file:str) -> None:

        ispdf = False
        if os.path.splitext(file)[1].lower() == '.pdf':
            ispdf = True

        if self.options['aim_pattern']:
            with open(self.options['aim_pattern'], mode='r', encoding='utf-8') as f:
                patterns = f.readlines()

            for p in patterns:
                if p.startswith('~~~') or p.startswith('```'):
                    continue
                p = re.sub('\t.*$', '', p)
                p = p.rstrip()
                if p not in self.found_count.keys():
                    self.found_count[p] = 0
                if ispdf:
                    self.found_count[p] += self.find_pdf(file, p)
                else:
                    self.found_count[p] += self.find_txt(file, p)
        else:
            if ispdf:
                self.find_pdf(file, self.options['aim'])
            else:
                self.find_txt(file, self.options['aim'])


    def find_txt(self, file:str, pattern:str) -> int:

        if self.options['dotall']:
            return self.find_txt_entire(file, pattern)
        else:
            return self.find_txt_byline(file, pattern)

    def find_txt_byline(self, file:str, pattern:str) -> int:

        count = 0
        found_line = ""

        try:
            with open(file, mode='r', encoding='utf-8') as f:
                content = f.readlines()
        except:
            print('{} is not encoded in UTF-8.'.format(file))
            return 0

        for num, line in enumerate(content):
            if self.options['case_sensitive']:
                matched = re.search(pattern, line)
            else:
                matched = re.search(pattern, line, flags=re.IGNORECASE)
            if matched:
                found_line = found_line + "  line {:4}: {}\n".format(num+1, line.strip())
                count += 1
        if count > 0:
            self.add_found(f"\x1b[32m{file}: {count}\x1b[0m")
            self.add_found(found_line)

        return count


    def find_txt_entire(self, file:str, pattern:str) -> int:

        count = 0
        found_line = ""

        try:
            with open(file, mode='r', encoding='utf-8') as f:
                content = f.read()
        except:
            print('{} is not encoded in UTF-8.'.format(file))
            return 0

        matched = re.findall(pattern, content, re.DOTALL)
        if matched:
            for i in range(len(matched)):
                found_line = found_line + "\t{}\n".format(matched[i])
                count += 1 

        if count > 0:
            self.add_found(f"\x1b[32m{file}: {count}\x1b[0m")
            self.add_found(found_line)

        return count


    def find_pdf(self, file:str, pattern:str) -> int:

        count = 0
        found_line = ""

        doc = pymupdf.open(file)
        for page_no in range(0, doc.page_count):
            page = doc.load_page(page_no)
            page_text = page.get_text()
            page_text = re.sub('-\n', '', page_text)
            lines = page_text.split('\n')
            for line in lines:
                if self.options['case_sensitive']:
                    matched = re.search(pattern, line)
                else:
                    matched = re.search(pattern, line, flags=re.IGNORECASE)
                if matched:
                    # self.add_found("\tPage {}: {}".format(page_no+1, line))
                    found_line = found_line + "  page {:3}: {}\n".format(page_no+1, line.strip())
                    count += 1

        if count > 0:
            self.add_found(f"\x1b[32m{file}: {count}\x1b[0m")
            self.add_found(found_line)
        return count


    def add_found(self, found_line:str) -> None:

            if self.options['output'] is None:
                print(found_line)
            else:
                self.found.append(found_line)


    def write_found(self) -> None:

        content = '\n'.join(self.found)
        content = content.replace('\x1b[32m', '')
        content = content.replace('\x1b[0m', '')
        output = self.determine_output('', output=self.options['output'])
        with open(output, mode='w', encoding='utf-8') as f:
            f.write(content)
        if not self.options['quietly']:
            print(f"{output} has been created or updated.")


    def extract(self, file:str) -> None:

        if os.path.splitext(file)[1].lower() == '.pdf':
            return

        if self.options['aim_pattern'] is not None:
            with open(self.options['aim_pattern'], mode='r', encoding='utf-8') as f:
                ptrn = [line.rstrip() for line in f]
        else:
            ptrn = [self.options['aim']]

        with open(file, mode='r', encoding='utf-8') as f:
            content = f.read()

        found = []
        for i in ptrn:
            if self.options['dotall']:
                found_lines = re.findall(i, content, flags=re.DOTALL)
            else:
                found_lines = re.findall(i, content, flags=re.MULTILINE)
            if found_lines:
                if self.options['gather']:
                    self.found += found_lines
                else:
                    found += found_lines

        if self.options['extract']:
            found = list(set(found))
            content = '\n'.join(sorted(found, key=str.lower))
            output = self.determine_output(file, output='_extracted')
            with open(output, mode='w', encoding='utf-8') as f:
                f.write(content)


    def replace_expand_scope(self, content, pattern) -> list:

        within_scope = False
        for i, line in enumerate(content):
            if within_scope:
                matched = re.search(pattern[3], line)
                if matched:
                    within_scope = False
                else:
                    content[i] = re.sub(pattern[1], pattern[2], line)
            else:
                matched = re.search(pattern[0], line)
                if matched:
                    within_scope = True
        return content


    def replace_expand_inline(self, content:list, pattern:list) -> list:

        for i, line in enumerate(content):
            phrases = re.findall(pattern[0], line)
            for phrase in phrases:
                aim = '{}{}{}'.format(pattern[3], phrase, pattern[4])
                substitute = re.sub(pattern[1], pattern[2], phrase)
                if len(pattern) == 7:
                    substitute = '{}{}{}'.format(pattern[5], substitute, pattern[6])
                else:
                    substitute = '{}{}{}'.format(pattern[3], substitute, pattern[4])
                line = line.replace(aim, substitute)
            content[i] = line

        return content


    def replace_expand_simple(self, content:list, pattern:list) -> list:

        for i, line in enumerate(content):
            if len(pattern) == 1:
                content[i] = re.sub(pattern[0], '', line)
            else:
                content[i] = re.sub(pattern[0], pattern[1], line)

        return content


    def replace_expand(self, file:str) -> str:

        try:
            with open(file, mode='r', encoding='utf-8') as f:
                content = f.readlines()
        except:
            print('{} is not encoded in UTF-8.'.format(file))
            return None

        if not os.path.exists(self.options['pattern']):
            return None

        ptrn_ext = os.path.splitext(self.options['pattern'])[1].lower()
        with open(self.options['pattern'], mode='r', encoding='utf-8') as ptrn:
            if ptrn_ext == '.tsv':
                patterns = csv.reader(ptrn, delimiter='\t')
            else:
                patterns = csv.reader(ptrn)

            for pattern in patterns:
                if len(pattern) == 0 or pattern[0].startswith('~~~') or pattern[0].startswith('```'):
                    continue
                elif len(pattern) == 4:
                    content = self.replace_expand_scope(content, pattern)
                elif len(pattern) == 5 or len(pattern) == 7:
                    content = self.replace_expand_inline(content, pattern)
                else:
                    content = self.replace_expand_simple(content, pattern)

        return ''.join(content)


    # scanline('...', ['\\\\footnote', '{', '}'])
    def scan_line(self, line:str, pattern:list) -> list:

        leading = pattern[0] + pattern[1]
        line_out = ''
        inside = []

        while True:
            found = re.search(leading, line)
            if found:
                line_out += line[:found.end()]
                line = line[found.end():]
                at = 0
                opening = 1
                while opening > 0:
                    if line[at] == pattern[1]:
                        opening += 1
                    if line[at] == pattern[2]:
                        opening += -1
                    at += 1
                inside.append(line[:at-1])
                line_out += line[:at]
                line = line[at:]
            else:
                break

        return inside


    def replace_once(self, file:str) -> str:

        try:
            with open(file, mode='r', encoding='utf-8') as f:
                content = f.read()
        except:
            print('{} is not encoded in UTF-8.'.format(file))
            return None

        if self.options['pattern'] is None:
            if self.options['dotall']:
                content = re.sub(self.options['aim'], self.options['substitute'], content, flags=re.DOTALL)
            else:
                content = re.sub(self.options['aim'], self.options['substitute'], content, flags=re.MULTILINE)
        else:
            ptrn_ext = os.path.splitext(self.options['pattern'])[1].lower()
            with open(self.options['pattern'], mode='r', encoding='utf-8') as ptrn:
                if ptrn_ext == '.tsv':
                    patterns = csv.reader(ptrn, delimiter='\t')
                else:
                    patterns = csv.reader(ptrn)
                DOTALL = self.options['dotall']
                for pattern in patterns:
                    if len(pattern) == 0:
                        continue
                    elif pattern[0].startswith('~~~') or pattern[0].startswith('```'): # use ~~~ or ``` to write comments in TSV
                        if pattern[0].endswith('DOTALL'):
                            DOTALL = True
                        else:
                            DOTALL = False
                        continue
                    else:
                        if DOTALL:
                            if len(pattern) == 1:
                                content = re.sub(pattern[0], '', content, flags=re.DOTALL)
                            else:
                                content = re.sub(pattern[0], pattern[1], content, flags=re.DOTALL)
                        else:
                            if len(pattern) == 1:
                                content = re.sub(pattern[0], '', content, flags=re.MULTILINE)
                            else:
                                # when replacing with quotation mark (") use \" in TSV
                                if pattern[1] == '\\"':  
                                    content = re.sub(pattern[0], '"', content, flags=re.MULTILINE)
                                else:
                                    content = re.sub(pattern[0], pattern[1], content, flags=re.MULTILINE)

        return content


    def replace(self, file:str) -> None:
        
        if self.options['flag'] == 'EXPAND':
            content = self.replace_expand(file)
        else:
            content = self.replace_once(file)

        if content is None:
            return

        output = self.determine_output(file, output='_replaced')
        with open(output, mode='w', encoding='utf-8') as f:
            f.write(content)


    def compare(self) -> None:

        if len(self.targets) != 2:
            print("Two files are required at least.")
            return

        if '*' in self.targets[0]:
            if self.options['recursive']:
                dir1 = os.path.dirname(self.targets[0])
                dir2 = os.path.dirname(self.targets[1])
                filePattern1 = os.path.basename(self.targets[0])
                subdirs = [x[0] for x in os.walk(dir1)]
                for subdir in subdirs:
                    fnpattern = os.path.join(subdir, filePattern1)
                    for file1 in glob.glob(fnpattern):
                        file2 = file1.replace(dir1, dir2)
                        if os.path.exists(file1) and os.path.exists(file2):
                            self.compare_files(file1, file2)
            else:
                dir2 = os.path.dirname(self.targets[1])
                for file1 in glob.glob(self.targets[0]):
                    filename = os.path.basename(file1)
                    file2 = os.path.join(dir2, filename)
                    if os.path.exists(file1) and os.path.exists(file2):
                        self.compare_files(file1, file2)
        else:
            self.compare_files(self.targets[0], self.targets[1])

        content = '\n'.join(self.found)
        if self.options['output'] is None:
            print(content)
        else:
            with open(self.options['output'], mode='w', encoding='utf-8') as f:
                f.write(content)


    def compare_files(self, file1:str, file2:str) -> None:

        self.found.append(f"{file1}\n{file2}")
        with open(file1, "r", encoding='utf-8') as f1, open(file2, "r", encoding='utf-8') as f2:
            for lineno, (line1, line2) in enumerate(zip(f1, f2), start=1):
                if line1 != line2:
                    self.found.append(("{:5}: {}\n{:5}: {}".format(lineno, line1.strip(), lineno, line2.strip())))


    def determine_output_indefinite(self, file:str, output=None) -> str:

        iFilename, iExt = os.path.splitext(os.path.basename(file))

        oDir = os.path.dirname(output)
        oFilename, oExt = os.path.splitext(os.path.basename(output))
        oPrefix = ''
        oSuffix = ''

        if oFilename.startswith('.') or oFilename == '':
            oExt = oFilename
            oFilename = iFilename
        if oExt == '' or oExt == '.':
            oExt = iExt 
        if oFilename.startswith('_'):
            oSuffix = oFilename
            oFilename = iFilename
        if oFilename.endswith('_'):
            oPrefix = oFilename
            oFilename = iFilename

        if len(oDir) != 0:
            if not os.path.exists(oDir):
                os.mkdir(oDir)

        oFile = '{}{}{}{}'.format(oPrefix, oFilename, oSuffix, oExt)
        output = os.path.join(oDir, oFile).replace('/','\\')

        if not self.options['overwrite']:
            if os.path.exists(output):
                counter = 0
                while os.path.exists(output):
                    counter += 1
                    oFile = '{}{}{}_{}{}'.format(oPrefix, oFilename, oSuffix, counter, oExt)
                    output = os.path.join(oDir, oFile).replace('/','\\')

        return output


    def determine_output(self, file:str, output=None) -> str:

        if self.options['output'] is None:
            if self.options['overwrite']:
                if len(file) > 0:
                    return file
                else:
                    return output
            else:
                return self.determine_output_indefinite(file=file, output=output)
        else:
            if self.options['overwrite']:
                return self.options['output']
            else:
                return self.determine_output_indefinite(file=file, output=self.options['output'])


    def check_if_pdf(self, file:str) -> str:

        filename, ext = os.path.splitext(file)
        if os.path.splitext(file)[1].lower() == '.pdf':
            doc = pymupdf.open(file)
            content = ""
            for p in range(0, doc.page_count):
                page = doc.load_page(p)
                content += page.get_text()
            output = os.path.splitext(file)[0] + '_from_pdf.txt'
            with open(output, mode='w', encoding='utf-8') as f:
                f.write(content)
            return output
        else:
            return file


    def count_words(self, file:str) -> None:

        # Spaces are not counted as a character.
        lines, spc_chars, chars, words = 0, 0, 0, 0

        file = self.check_if_pdf(file)
        f = open(file, mode='r', encoding='utf-8')
        for line in f.readlines():
            lines += 1
            spc_chars += len(line)
            chars += len(line.replace(' ', ''))
            this = line.split(None)
            words += len(this)
        f.close()

        output = '''{}
 Lines: {:,}
 Words: {:,}
 Characters including spaces: {:,}
 Characters without spaces: {:,}'''.format(file, lines, words, spc_chars, chars) 
        print(output)

        self.lines += lines
        self.words += words
        self.spc_chars += spc_chars
        self.chars += chars
        self.files += 1


    def count_pdf_pages(self, file:str) -> None:

        doc = pymupdf.open(file)
        print('{}: {}'.format(file, doc.page_count))
        self.pages += doc.page_count


    def align_string(self, string: str, width: int) -> None:

        nfc_string = unicodedata.normalize('NFC', string)
        wide_chars = [unicodedata.east_asian_width(c) for c in nfc_string]
        num_wide_chars = sum(map(wide_chars.count, ['W', 'F']))
        width = max(width-num_wide_chars, num_wide_chars)
        return '{:{w}}'.format(nfc_string, w=width)


    def write_gathered(self) -> None:
        
        if len(self.found) < 1:
            print("There is nothing to write.")
            return

        # remove duplicates and sort
        self.found = list(set(self.found))
        content = '\n'.join(sorted(self.found, key=str.lower))
        output = self.determine_output('', output='gathered_strings.txt')
        with open(output, mode='w', encoding='utf-8') as f:
            f.write(content)
        if not self.options['quietly']:
            print(f"{output} has been created or updated.")


    def tsv_to_xlsx(self, file:str) -> None:

        output = self.determine_output_indefinite(file=file, output='.xlsx')

        content = []
        with open(file, mode='r', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter='\t')
            for row in reader:
                if self.options['escape_tex']: 
                    content.append(self.remove_tex(row))
                else:
                    content.append(row)

        workbook = openpyxl.Workbook()
        ws = workbook.active
        for row, line in enumerate(content):
            for column, text in enumerate(line):
                ws.cell(row=row+1, column=column+1, value=text)
        workbook.save(output)
        if not self.options['quietly']:
            print("{} -> {}".format(file, output))


    def columns_to_remove(self, columns_to_extract:str, max_col:int) -> list:

        entire_columns = list(range(1,max_col+1))
        include_columns = []
        exclude_columns = []
        exclude_column_ranges = []

        # 1, 3-5 -> 1, 3, 4, 5
        incol = columns_to_extract.split(',')
        for v in incol:
            if '-' in v:
                column_range = v.split('-')
                lower = int(column_range[0])
                upper = int(column_range[1]) + 1
                include_columns += list(range(lower, upper))
            else:
                include_columns.append(int(v))
        include_columns = sorted(set(include_columns))

        # 2, 6, 7, ... -> [2, 2], [5, n]
        for v in entire_columns:
            if not v in include_columns:
                exclude_columns.append(v)

        lower = exclude_columns[0]
        upper = exclude_columns[0]
        i = 1
        while i < len(exclude_columns): 
            if exclude_columns[i] > upper + 1:
                exclude_column_ranges += [[lower, upper]]
                lower = exclude_columns[i]
                upper = exclude_columns[i]
            else:
                upper = exclude_columns[i]
            i += 1
        exclude_column_ranges += [[lower, upper]]
        exclude_column_ranges = exclude_column_ranges[::-1]

        if not self.options['quietly']:
            print(exclude_column_ranges)
        return exclude_column_ranges

    # C:\>wordig -t -a "1,3-5" -o goo.tsv foo.xlsx
    def xlsx_to_tsv(self, xlsxFile:str) -> None:

        if PurePath(xlsxFile).suffix.lower() != '.xlsx':
            print('xlsx 파일을 지정하시오.')
            return

        self.options['overwrite'] = True
        fileName = os.path.splitext(os.path.basename(xlsxFile))[0]
        if not self.options['quietly']:
            print(xlsxFile)

        workbook = openpyxl.load_workbook(xlsxFile)
        sheets = workbook.sheetnames

        for i in range(len(sheets)):
            sheet = workbook[sheets[i]]
            sheetName = sheets[i]
            output = f"{fileName}_{sheetName}.tsv"
            output = self.determine_output_indefinite(xlsxFile, output)
            self.sheet_to_tsv(sheet, output)

    def sheet_to_tsv(self, worksheet, output):

        if self.options['aim'] is not None:
            column_ranges = self.columns_to_remove(self.options['aim'], worksheet.max_column)
            for i in column_ranges:
                worksheet.delete_cols(i[0], i[1]-i[0]+1)

        line = []
        content = ''
        for row in worksheet.iter_rows():
            line.clear()
            for cell in row:
                if isinstance(cell.value, str):
                    line.append(cell.value)
                elif (isinstance(cell.value, int) or isinstance(cell.value, float)):
                    line.append(str(cell.value))
                else:
                    line.append(' ')
            if len(line) > 0 :
                tmp = ''.join(line)
                if tmp.strip() != '': # 공백을 제거한 뒤에 빈 줄이 아니라면
                    content += self.escape_tex('\t'.join(line))

        with open(output, mode='w', encoding='utf-8') as f:
            f.write(content)
        if not self.options['quietly']:
            print(f"-> {output}")


        # if os.path.splitext(output)[1].lower() == '.xlsx':
        #     workbook.save(filename=output)
        #     if not self.options['quietly']:
        #         print("{} -> {}".format(xlsxFile, output))
        #     return
        # else:
        #     line = []
        #     content = ''
        #     for row in worksheet.iter_rows():
        #         line.clear()
        #         for cell in row:
        #             if isinstance(cell.value, str):
        #                 line.append(cell.value)
        #             elif (isinstance(cell.value, int) or isinstance(cell.value, float)):
        #                 line.append(str(cell.value))
        #             else:
        #                 line.append(' ')
        #         if len(line) > 0 :
        #             tmp = ''.join(line)
        #             if tmp.strip() != '': # 공백을 제거한 뒤에 빈 줄이 아니라면
        #                 content += self.escape_tex('\t'.join(line))

        #     with open(output, mode='w', encoding='utf-8') as f:
        #         f.write(content)
        #     if not self.options['quietly']:
        #         print("{} -> {}".format(xlsxFile, output))


    def escape_tex(self, string:str) -> str:

        if self.options['escape_tex']: 
            string = re.sub(r'\\', '\\\\textbackslash', string)
            string = re.sub('([&%$#_])', '\\\\\\1', string)
            string = re.sub('\n', '\\\\linebreak{}', string)
        string += '\n'
        return string


    def remove_tex(self, columns: list) -> list:

        for i, string in enumerate(columns):
            string = re.sub('\\\\textbackslash', '', string)
            string = re.sub('\\\\([&%$#_])', '\\1', string)
            string = re.sub('\\\\linebreak\\{\\}', '\\n', string)
            columns[i] = string
        return columns


    def beautifyML(self, filePath:str) -> None:

        fileExtension = os.path.splitext(filePath)[1].lower()        
        if fileExtension == '.xml':
            print(filePath)
            self.XmlFormatter(filePath)
        elif fileExtension == '.html':
            print(filePath)
            self.HtmlFormatter(filePath)


    def HtmlFormatter(self, htmlFile:str) -> None:

        if Path(htmlFile).stat().st_size == 0:
            print("It's empty.")    
            return

        with open(htmlFile, mode='r', encoding='utf-8') as fs:
            content = fs.read()
        soup = bs4.BeautifulSoup(content, 'html.parser')
        htmlFormatter = bs4.formatter.HTMLFormatter(indent=int(self.options['indent']))
        content = soup.prettify(formatter=htmlFormatter)
        with open(htmlFile, mode='w', encoding='utf-8') as fs:
            fs.write(content)


    def XmlFormatter(self, xmlFile:str) -> None:

        if Path(xmlFile).stat().st_size == 0:
            print("It's empty.")    
            return
        
        with open(xmlFile, mode='r', encoding='utf-8') as fs:
            content = fs.read()
        tree = ET.fromstring(content)
        ET.indent(tree, space='    ')
        content = ET.tostring(tree, encoding='unicode')
        content = content.replace("\u2028", "&#x2028;")
        content = content.replace("\u2029", "&#x2029;")
        with open(xmlFile, mode='w', encoding='utf-8') as fs:
            fs.write(content)


    def determine_task(self) -> None:

        if self.options['unicode']:
            UTF = UnicodeDigger(chars=self.targets[0])
            UTF.print()
        elif self.options['unicode_bits']:
            UTF = UnicodeDigger(chars=self.targets[0], flag=1)
            UTF.print()
        elif self.options['unicode_hexadecimal']:
            UnicodeDigger.char(self.targets)
        elif self.options['unicode_decimal']:
            UnicodeDigger.char(self.targets, False)

        elif self.options['xlsx']:
            self.run_recursive(self.tsv_to_xlsx)
        elif self.options['tsv']:
            self.run_recursive(self.xlsx_to_tsv)
        elif self.options['compare']:
            self.compare()
        elif self.options['pattern'] is not None:
            if os.path.exists(self.options['pattern']):
                self.run_recursive(self.replace)
            else:
                print('{} is not found.'.format(self.options['pattern']))
        elif self.options['aim_pattern'] is not None:
            if os.path.exists(self.options['aim_pattern']):
                if self.options['extract'] or self.options['gather']:
                    self.run_recursive(self.extract)
                    if self.options['gather']:
                        self.write_gathered()
                else:
                    self.run_recursive(self.find)
                    for key in self.found_count.keys():
                        print('{}: {}'.format(key, self.found_count[key]))
                    if self.options['output']:
                        self.write_found()
            else:
                print('{} is not found.'.format(self.options['aim_pattern']))
        elif self.options['aim']:
            if self.options['substitute'] is not None:
                self.run_recursive(self.replace)
            else:
                if self.options['extract'] or self.options['gather']:
                    self.run_recursive(self.extract)
                    if self.options['gather']:
                        self.write_gathered()
                else:
                    self.run_recursive(self.find)
                    if self.options['output']:
                        self.write_found()
        elif self.options['page_count']:
            self.run_recursive(self.count_pdf_pages)
            print( 'Total pages: {:,}'.format(self.pages) )
        elif self.options['beautify']:
            self.run_recursive(self.beautifyML)
        else:
            self.run_recursive(self.count_words)
            if self.files > 1:
                output = '''Total
 Lines: {:,}
 Words: {:,} 
 Characters including spaces: {:,}
 Characters without spaces: {:,}'''.format(self.lines, self.words, self.spc_chars, self.chars) 
                print(output)



class UnicodeDigger(object):


    def __init__(self, chars=None, flag=0):

        self.chars = chars
        self.flag = flag


    def highlight_binary_code(self, dec, byte) -> str:

        # 31:red, 32:green, 33:yellow, 34:blue, 35:magenta, 36:cyan, 37: white
        head = '\x1b[37m'
        tail = '\x1b[36m'
        normal = '\x1b[0m'

        if dec < int('0x80', 16):
            byte = byte.zfill(8)
            return head + byte[:1] + tail + byte[1:] + normal
        elif dec < int('0x800', 16):
            byte = byte.zfill(12)
            return head + byte[0:6] + tail + byte[6:] + normal
        elif dec < int('0x10000', 16):
            byte = byte.zfill(16)
            return head + byte[0:4] + tail + byte[4:10] + head + byte[10:16] + normal
        else:
            byte = byte.zfill(24)
            return head + byte[0:6] + tail + byte[6:12] + head + byte[12:18] + tail + byte[18:24] + normal


    def highlight_binary_byte(self, byte_number, byte_index, byte) -> str:

        head = '\x1b[32m'
        tail = '\x1b[33m'
        normal = '\x1b[0m'

        if byte_index > 0:
            return head + byte[:2] + tail + byte[2:] + normal
        else:
            if byte_number == 2:
                return head + byte[:3] + tail + byte[3:] + normal
            elif byte_number == 3:
                return head + byte[:4] + tail + byte[4:] + normal
            elif byte_number == 4:
                return head + byte[:5] + tail + byte[5:] + normal


    def print(self) -> None:

        for char in self.chars:
            charname = unicodedata.name(char).lower()
            # decimal code points
            Dcode = ord(char)
            # hexadecimal code points
            Hcode = hex(Dcode).upper().replace('0X', '0x')
            # binary code points
            Bcode = bin(Dcode).replace('0b', '')
            Bcode = self.highlight_binary_code(Dcode, Bcode)

            # hexadecimal UTF-8 bytes
            if Dcode > 127:
                Hbyte = str(char.encode('utf-8'))
                Hbyte = Hbyte.replace("b'\\x", "")
                Hbyte = Hbyte.replace("'", "")
                Hbyte = Hbyte.split('\\x')
                # binary UTF-8 bytes
                Bbyte = []
                for i, val in enumerate(Hbyte):
                    Hbyte[i] = val.upper()
                    bbyte = bin(int(val, 16)).replace('0b', '')
                    bbyte = self.highlight_binary_byte(len(Hbyte), i, bbyte)
                    Bbyte.append(bbyte)
                Hbyte = ''.join(Hbyte)
                Bbyte = ' '.join(Bbyte)
                if self.flag == 1:
                    print(char, Dcode, Hcode, Bcode, Hbyte, Bbyte, charname)
                else:
                    print(char, Dcode, Hcode, charname)
            else:
                if self.flag == 1:
                    print(char, Dcode, Hcode, Bcode, charname)
                else:
                    print(char, Dcode, Hcode, charname)


    def char(codepoints, hex=True) -> None:

        for i in codepoints:
            if hex:
                codepoint = int(i, 16)
                if codepoint >= 0xE000 and codepoint <= 0xF8FF:
                    print('Private Use Area')
                else:
                    try:
                        char = chr(codepoint)
                        charname = unicodedata.name(char).lower()
                        print(char, charname) #end=' '
                    except:
                        print('Enter hexadecimal numbers greater than x19.')
            else:
                codepoint = int(i)
                if codepoint >= 57344 and codepoint <= 63743:
                    print('Private Use Area')
                else:
                    try:
                        char = chr(codepoint)
                        charname = unicodedata.name(char).lower()
                        print(char, charname) #end=' '
                    except:
                        print('Enter decimal numbers greater than 31.')


def parse_args() -> argparse.Namespace:

    parser=argparse.ArgumentParser(
        description="Count words, count pages, find or replace strings, and more."
    )

    parser.add_argument(
        'targets', nargs='+',
        help='Specify one or more text files or characters.'
    )
    parser.add_argument(
        '-a', dest='aim', default=None,
        help='Specify a string to find.'
    )
    parser.add_argument(
        '-A', dest='aim_pattern', default=None,
        help='Specify a text file which contains regular expressions for text search.'
    )
    parser.add_argument(
        '-s', dest='substitute', default=None,
        help='Specify a string with which to replace found strings.'
    )
    parser.add_argument(
        '-P', dest='pattern', default=None,
        help='Specify a TSV or CSV file which contains regular expressions for text replacement.'
    )
    parser.add_argument(
        '-C', '--compare', action='store_true', default=False,
        help='Specify two files or folders to compare them.'
    )
    parser.add_argument(
        '-c', '--case-sensitive', dest='case_sensitive', action='store_true', default=False,
        help='Perform case-sensitive match when finding. Replacing is always case-sensitive.'
    )
    parser.add_argument(
        '-L', '--dotall', dest='dotall', action='store_true', default=False,
        help='Dot matches all characters including newline.'
    )
    parser.add_argument(
        '-F', '--flag', dest='flag', default='ONCE',
        help='Specify "EXPAND" or "2" to perform in-scope substitutions.'
    )
    parser.add_argument(
        '-e', '--extract', dest='extract', action='store_true', default=False,
        help='Extract strings. Use with -a or -A option.'
    )
    parser.add_argument(
        '-g', '--gather', dest='gather', action='store_true', default=False,
        help='Extract and gather strings. Use with -a or -A option.'
    )
    parser.add_argument(
        '-o', dest='output', default=None, 
        help='Specify a directory or filename for output.'
    )
    parser.add_argument(
        '-v', '--overwrite', dest='overwrite', action='store_true', default=False,
        help='Overwrite the source file when replacing.'
    )
    parser.add_argument(
        '-r', '--recursive', dest='recursive', action='store_true', default=False,
        help='Search through all subdirectories.'
    )
    parser.add_argument(
        '-p', '--count-pdf-page', dest='page_count', action='store_true', default=False,
        help='Count PDF pages.'
    )
    parser.add_argument(
        '-U', '--unicode-info', dest='unicode', action='store_true', default=False,
        help='Show the unicode information of a given character.',
    )
    parser.add_argument(
        '-u', '--unicode-utf8', dest='unicode_bits', action='store_true', default=False,
        help='Show the unicode information with its UTF-8 bits.',
    )
    parser.add_argument(
        '-X', '--hexadecimal', dest='unicode_hexadecimal', action='store_true', default=False,
        help='Show the unicode character of a given hexadecimal.',
    )
    parser.add_argument(
        '-D', '--decimal', dest='unicode_decimal', action='store_true', default=False,
        help='Show the unicode character of a given decimal.',
    )
    parser.add_argument(
        '-x', '--tsv-to-xlsx', dest='xlsx', action='store_true', default=False,
        help='Specify a TSV file or more from which to convert to XLSX.'
    )
    parser.add_argument(
        '-t', '--xlsx-to-tsv', dest='tsv', action='store_true', default=False,
        help='Specify a XLSX file or more from which to convert to TSV or XLSX.'
    )
    parser.add_argument(
        '-T', '--escape-tex', dest='escape_tex', action='store_true', default=False,
        help='Escape TeX macros when converting between XLSX and TSV.'
    )
    parser.add_argument(
        '-q', '--quietly', dest='quietly', action='store_true', default=False,
        help='Display no result message.'
    )
    parser.add_argument(
        '-b', '--beautiyfy', dest='beautify', action='store_true', default=False,
        help='Beautify HTML or XML files.'
    )
    parser.add_argument(
        '-i', dest='indent', default=4, 
        help='Specify a number to change the indent size.'
    )
    

    return parser.parse_args()


if __name__ == '__main__':
    args = parse_args()
    options = vars(args)
    targets = options['targets']
    del options['targets']
    WordDigger(targets, options)